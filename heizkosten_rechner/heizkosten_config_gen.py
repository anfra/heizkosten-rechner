"""
heizkosten_config_gen.py – Konfigurationsdatei für die Heizkostenabrechnung generieren.

Liest Objekt- und Mieterdaten aus dem immocloud-Objektexport (XLSX) und erzeugt
eine YAML-Konfigurationsdatei, die von der Berechnungs-App verwendet wird.

Energiekosten werden mit Dummy-Werten vorbelegt und müssen manuell ergänzt werden.

Verwendung:
    python heizkosten_config_gen.py <export.xlsx> [--jahr 2025] [--out heizkosten_config.yaml]

Abhängigkeiten:
    pip install openpyxl pyyaml
"""

import argparse
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
import yaml


# ---------------------------------------------------------------------------
# Immocloud-Einheitenname → interne NE-Nummer (aus UNIT_MAP)
# ---------------------------------------------------------------------------

EINHEIT_TO_NR = {
    "10 - EG rechts":    "2",
    "20 - EG links":     "1",
    "30 - 1. OG rechts": "4",
    "40 - 1. OG links":  "3",
    "50 - 2. OG rechts": "6",
    "60 - 2. OG links":  "5",
    "70 - DG rechts":    "8",
    "80 - DG links":     "7",
}


# ---------------------------------------------------------------------------
# XLSX lesen
# ---------------------------------------------------------------------------

def parse_date(val) -> Optional[str]:
    """Wandelt Datumswert aus Excel (str 'TT.MM.JJJJ' oder datetime) in ISO-Format."""
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def load_xlsx(path: Path, abrechnungsjahr: int) -> dict:
    """
    Liest den immocloud-Objektexport und gibt strukturierte Daten zurück.

    Gibt zurück:
        {
          "objekt": {"adresse": ..., "strasse": ..., "plz": ..., "ort": ...},
          "einheiten": {
            "1": {
              "name": "20 - EG links",
              "flaeche_m2": 50.4,
              "mieter": [
                {"name": "...", "einzug": "YYYY-MM-DD", "auszug": "YYYY-MM-DD" | null}
              ]
            },
            ...
          }
        }
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Kopfzeile finden (erste Zeile mit "Einheit" als Zellinhalt)
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and "Einheit" in row:
            header_row = i
            headers = list(row)
            break

    if header_row is None:
        raise ValueError("Keine Kopfzeile mit 'Einheit' in der Excel-Datei gefunden.")

    col = {name: idx for idx, name in enumerate(headers) if name}

    objekt_info = {}
    einheiten: dict = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue

        einheit_name = row[col.get("Einheit", -1)] if "Einheit" in col else None
        if not einheit_name or not isinstance(einheit_name, str):
            continue

        einheit_name = einheit_name.strip()
        ne_nr = EINHEIT_TO_NR.get(einheit_name)
        if ne_nr is None:
            print(f"  Warnung: Unbekannte Einheit '{einheit_name}' – übersprungen.")
            continue

        # Objekt einmalig erfassen
        if not objekt_info and "Objekt" in col:
            objekt_info = {
                "adresse":   str(row[col["Objekt"]] or "").strip(),
                "strasse":   str(row[col.get("Straße", col["Objekt"])] or "").strip(),
                "hausnummer": str(row[col.get("Hausnummer", 0)] or "").strip(),
                "plz":       str(row[col.get("PLZ", 0)] or "").strip(),
                "ort":       str(row[col.get("Ort", 0)] or "").strip(),
            }

        # Wohnfläche: erste nicht-None-Angabe je Einheit verwenden
        flaeche = row[col["Wohnfläche"]] if "Wohnfläche" in col else None
        mietername = str(row[col["Mietername"]] or "").strip() if "Mietername" in col else ""
        einzug_raw = row[col["Mietbeginn"]] if "Mietbeginn" in col else None
        auszug_raw = row[col["Mietende"]]   if "Mietende"  in col else None

        einzug = parse_date(einzug_raw)
        auszug = parse_date(auszug_raw)

        # Auszug auf Jahresende kürzen: wenn nach dem Abrechnungsjahr, ignorieren
        if auszug:
            auszug_dt = datetime.strptime(auszug, "%Y-%m-%d")
            if auszug_dt.year > abrechnungsjahr:
                auszug = None  # noch aktiver Mieter im Abrechnungsjahr
            elif auszug_dt.year < abrechnungsjahr:
                continue  # Mieter vor dem Abrechnungsjahr ausgezogen → irrelevant

        # Einzug: wenn noch nicht im Abrechnungsjahr, als Jahresanfang behandeln
        aktiv_im_jahr = False
        if einzug:
            einzug_dt = datetime.strptime(einzug, "%Y-%m-%d")
            if einzug_dt.year <= abrechnungsjahr:
                aktiv_im_jahr = True
                if einzug_dt.year < abrechnungsjahr:
                    einzug = f"{abrechnungsjahr}-01-01"  # normiert auf Jahresanfang
        if not aktiv_im_jahr:
            continue

        # Einheit anlegen oder ergänzen
        if ne_nr not in einheiten:
            einheiten[ne_nr] = {
                "name":       einheit_name,
                "flaeche_m2": float(flaeche) if isinstance(flaeche, (int, float)) and flaeche else None,
                "mieter":     [],
            }
        elif flaeche and isinstance(flaeche, (int, float)) and einheiten[ne_nr]["flaeche_m2"] is None:
            einheiten[ne_nr]["flaeche_m2"] = float(flaeche)

        if mietername:
            mieter_entry = {"name": mietername, "einzug": einzug, "auszug": auszug}
            # Duplikate vermeiden
            if mieter_entry not in einheiten[ne_nr]["mieter"]:
                einheiten[ne_nr]["mieter"].append(mieter_entry)

    return {"objekt": objekt_info, "einheiten": einheiten}


# ---------------------------------------------------------------------------
# YAML-Konfiguration generieren
# ---------------------------------------------------------------------------

def generate_config(data: dict, abrechnungsjahr: int, csv_datei: str) -> dict:
    """Baut die vollständige Konfigurationsstruktur für die Heizkostenabrechnung."""
    obj = data["objekt"]
    einheiten = data["einheiten"]

    # Gesamtfläche
    flaechen = [e["flaeche_m2"] for e in einheiten.values() if e["flaeche_m2"]]
    gesamtflaeche = round(sum(flaechen), 2)

    config = {
        "# HINWEIS": (
            "Energiekosten sind mit Dummy-Werten vorbelegt. "
            "Bitte mit tatsächlichen Werten aus der Versorger-Rechnung ersetzen."
        ),

        "objekt": {
            "adresse":         f"{obj.get('strasse', '')} {obj.get('hausnummer', '')}".strip(),
            "plz":             obj.get("plz", ""),
            "ort":             obj.get("ort", ""),
            "abrechnungsjahr": abrechnungsjahr,
            "csv_datei":       csv_datei,
            "gesamtflaeche_m2": gesamtflaeche,
        },

        "nutzeinheiten": {
            nr: {
                "name":       e["name"],
                "flaeche_m2": e["flaeche_m2"],
                "mieter":     e["mieter"],
            }
            for nr, e in sorted(einheiten.items(), key=lambda x: int(x[0]))
        },

        "kosten": {
            "# Heizung": "Gesamtkosten Heizenergie inkl. CO₂-Abgabe aus Versorger-Rechnung",
            "heizung_gesamt":        0.00,   # EUR – PFLICHTFELD: bitte ergänzen
            "co2_abgabe":            0.00,   # EUR – Anteil CO₂-Abgabe (in heizung_gesamt enthalten)

            "# Warmwasser": "Energieanteil Warmwasserbereitung + Kaltwasserkosten",
            "warmwasser_waerme":     0.00,   # EUR – aus WMZ-Ablesung oder § 9a Schätzung
            "warmwasser_wasser":     0.00,   # EUR – Kaltwasserpreis × Warmwasserverbrauch m³

            "# CO2": "Spezifischer CO₂-Ausstoß des Gebäudes (bestimmt Vermieter/Mieter-Split)",
            "co2_spezifisch_kg_m2":  0.0,   # kg CO₂/m²/Jahr – aus Energieausweis oder Rechnung
        },

        "aufteilung": {
            "# HeizkostenV § 6/§ 8": "30–50 % Grundkosten, 50–70 % Verbrauchskosten",
            "heizung_grundkosten_pct":    30,
            "warmwasser_grundkosten_pct": 30,
        },
    }

    return config


# ---------------------------------------------------------------------------
# YAML-Ausgabe (mit Kommentaren)
# ---------------------------------------------------------------------------

class CommentedDumper(yaml.Dumper):
    """YAML-Dumper mit angepasster Formatierung."""
    pass


def str_representer(dumper, data):
    if "\n" in data:
        return dumper.represent_scalar("tag:yaml.org,2002:str", data, style="|")
    return dumper.represent_scalar("tag:yaml.org,2002:str", data)


CommentedDumper.add_representer(str, str_representer)


def write_config(config: dict, out_path: Path) -> None:
    """Schreibt die Konfiguration als YAML-Datei mit erklärenden Kommentaren."""

    # Kommentar-Keys aus dem Dict herauslösen (werden als YAML-Kommentare geschrieben)
    # Einfachster Weg: manuell aufbauen statt yaml.dump der ganzen Struktur

    lines = [
        f"# Heizkostenkonfiguration — automatisch generiert am {date.today().isoformat()}",
        f"# Quelle: immocloud Objektexport",
        f"# ACHTUNG: Abschnitt 'kosten' mit echten Werten aus der Versorger-Rechnung füllen!",
        "",
    ]

    obj = config["objekt"]
    lines += [
        "objekt:",
        f"  adresse:          \"{obj['adresse']}\"",
        f"  plz:              \"{obj['plz']}\"",
        f"  ort:              \"{obj['ort']}\"",
        f"  abrechnungsjahr:  {obj['abrechnungsjahr']}",
        f"  csv_datei:        \"{obj['csv_datei']}\"",
        f"  gesamtflaeche_m2: {obj['gesamtflaeche_m2']}",
        "",
    ]

    lines.append("nutzeinheiten:")
    for nr, ne in config["nutzeinheiten"].items():
        flaeche = ne["flaeche_m2"] if ne["flaeche_m2"] is not None else "# FEHLEND – bitte ergänzen"
        lines.append(f"  \"{nr}\":")
        lines.append(f"    name:       \"{ne['name']}\"")
        lines.append(f"    flaeche_m2: {flaeche}")
        lines.append(f"    mieter:")
        for m in ne["mieter"]:
            auszug = f"\"{m['auszug']}\"" if m["auszug"] else "null"
            lines.append(f"      - name:   \"{m['name']}\"")
            lines.append(f"        einzug: \"{m['einzug']}\"")
            lines.append(f"        auszug: {auszug}")
    lines.append("")

    lines += [
        "# ─────────────────────────────────────────────────────────────────────",
        "# ENERGIEKOSTEN — bitte aus der Versorger-Jahresrechnung eintragen",
        "# ─────────────────────────────────────────────────────────────────────",
        "kosten:",
        "  # Gesamtkosten Heizenergie (Brennstoff + Wartung + CO₂-Abgabe)",
        "  heizung_gesamt:        0.00   # EUR — PFLICHTFELD",
        "  co2_abgabe:            0.00   # EUR — Anteil CO₂-Abgabe (in heizung_gesamt enthalten)",
        "",
        "  # Warmwasserbereitung",
        "  warmwasser_waerme:     0.00   # EUR — aus WMZ oder § 9a HeizkostenV",
        "  warmwasser_wasser:     0.00   # EUR — Kaltwasserpreis × WW-Verbrauch m³",
        "",
        "  # CO₂-Kosten (CO₂KostAufG) — spezifischer Ausstoß für Stufentabelle",
        "  co2_spezifisch_kg_m2:  0.0    # kg CO₂/m²/Jahr — aus Energieausweis",
        "",
        "# ─────────────────────────────────────────────────────────────────────",
        "# AUFTEILUNGSSCHLÜSSEL (§ 6 / § 8 HeizkostenV, 30–50 % Grundkosten)",
        "# ─────────────────────────────────────────────────────────────────────",
        "aufteilung:",
        "  heizung_grundkosten_pct:    30   # Anteil Grundkosten Heizung in %",
        "  warmwasser_grundkosten_pct: 30   # Anteil Grundkosten Warmwasser in %",
    ]

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"\n✓ Konfiguration geschrieben: {out_path}")


# ---------------------------------------------------------------------------
# Hauptprogramm
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Heizkostenkonfiguration aus immocloud-Objektexport (XLSX) generieren."
    )
    parser.add_argument("xlsx", help="Pfad zur immocloud-Exportdatei (.xlsx)")
    parser.add_argument("--jahr",  type=int, default=2025, help="Abrechnungsjahr (Standard: 2025)")
    parser.add_argument("--csv",   default=None,
                        help="Pfad zur Ablesewerte-CSV (optional, f\u00fcr sp\u00e4tere Erweiterung)")
    parser.add_argument("--out",   default="heizkosten_config.yaml",
                        help="Ausgabedatei (Standard: heizkosten_config.yaml)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path  = Path(args.out)

    if not xlsx_path.exists():
        print(f"Fehler: Datei nicht gefunden: {xlsx_path}")
        sys.exit(1)

    print(f"Lese {xlsx_path} ...")
    data = load_xlsx(xlsx_path, args.jahr)

    print(f"\nObjekt:    {data['objekt'].get('adresse', '')} {data['objekt'].get('ort', '')}")
    print(f"Einheiten: {len(data['einheiten'])} gefunden")
    for nr, e in sorted(data["einheiten"].items(), key=lambda x: int(x[0])):
        m_namen = ", ".join(m["name"] for m in e["mieter"]) or "—"
        flaeche = f"{e['flaeche_m2']} m²" if e["flaeche_m2"] else "FEHLEND"
        print(f"  NE {nr}: {e['name']:20s}  {flaeche:10s}  Mieter: {m_namen}")

    config = generate_config(data, args.jahr, args.csv)
    write_config(config, out_path)

    print("\nNächste Schritte:")
    print(f"  1. {out_path} öffnen")
    print(f"  2. Abschnitt 'kosten' mit Werten aus der Versorger-Rechnung füllen")
    print(f"  3. python heizkosten_calc.py {out_path}")


if __name__ == "__main__":
    main()
